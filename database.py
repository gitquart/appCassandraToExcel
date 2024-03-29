from cassandra.cluster import Cluster
from cassandra.auth import PlainTextAuthProvider
from cassandra.query import SimpleStatement
from openpyxl import Workbook
from openpyxl import load_workbook
from InternalControl import cInternalControl

objControl= cInternalControl()
cloud_config= {
        'secure_connect_bundle':'secure-connect-'+objControl.db+'_serverless.zip'
    }


def getCluster():
    #Connect to Cassandra
    objCC=CassandraConnection()
    user=''
    password=''
    if objControl.db=='dbquart':
        user=objCC.cc_user_dbquart
        password=objCC.cc_pwd_dbquart
    else:
        user=objCC.cc_user_test
        password=objCC.cc_pwd_test

    auth_provider = PlainTextAuthProvider(user,password)
    cluster = Cluster(cloud=cloud_config, auth_provider=auth_provider)

    return cluster

#getLargeQueryAndPrintToExcel
#Prints normal query in an spreadsheet
def getLargeQueryAndPrintToExcel(query,dir_excel,title):
    cluster = getCluster()
    session = cluster.connect()
    session.default_timeout=70     
    statement = SimpleStatement(query, fetch_size=1000)
    wb = load_workbook(dir_excel)
    ws = wb[title]
    
    for row in session.execute(statement):
        ls=[]
        for col in row:
            ls.append(str(col))
        ws.append(ls)
           
    wb.save(dir_excel) 
    cluster.shutdown() 

#getLargeQueryAndPrintToExcel_Special
#Prints a query in spreadsheet by special conditions
def getLargeQueryAndPrintToExcel_Special(query,dir_excel,title):
    cluster = getCluster()
    session = cluster.connect()
    session.default_timeout=70     
    statement = SimpleStatement(query, fetch_size=1000)
    wb = load_workbook(dir_excel)
    ws = wb[title]
    
    for row in session.execute(statement):
        ls=[]
        coln=1
        for col in row:
            #Case for cip , position 1
            if coln==1:
                if col!='':
                    chunks=str(col).split(';')
                    if len(chunks)>0:
                        lsPart=[]
                        for item in chunks:
                            strParts=item.strip()
                            lsPart.append(strParts[0])
                        col=';'.join(lsPart)    
                    else:
                        strParts=col.split()
                        col=strParts[0]        
            ls.append(str(col))
            coln+=1
        ws.append(ls)
           
    wb.save(dir_excel) 
    cluster.shutdown() 


def getLargeQuery(query):
    cluster = getCluster()
    session = cluster.connect()
    session.default_timeout=70     
    statement = SimpleStatement(query, fetch_size=1000)
    count=0
    print('Start count...')
    for row in session.execute(statement):
        count+=1
        
    print('Total rows:',str(count))        
    cluster.shutdown()     



def getShortQuery(query):
    res=''
    cluster=getCluster()
    session = cluster.connect()
    session.default_timeout=70
    #Check wheter or not the record exists      
    future = session.execute_async(query)
    res=future.result()
    cluster.shutdown()

    return res 


                

     
class CassandraConnection():
    cc_user_test='MpvtYRWPigKTDLxfcZMNIfYQ'
    cc_pwd_test='joFIPwAoFL_JWNsgAr,xTNf30gZpZu3Fg6,eMxQjKxm-Gz5Uva2R.ELwwzj88f4XPePGhXLWU89xzzP8a1BdgSpwLN+iPHhQpjRmXnvA-cbmvoK_In8Sr,MGadqF+TAh'
    cc_user_dbquart='psXzplCMoTnTjWLSeAblXSkr'
    cc_pwd_dbquart='_vGwtwEUsOxj9ZuSbm6SOJhyQOsU-s9QEzX7ZWZcRYKeew9sOqPeFmKr-pDzsIoisiF8aInS1HjSs7_fqZ9EZaMS96TNCwn_yReyjZvHuys+_fGSBuHKZz_+NTj6Z6L0'