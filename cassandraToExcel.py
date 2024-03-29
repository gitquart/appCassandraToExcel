"""
Program that will write cassandra table in excel workbook

"""
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import database as bd
from InternalControl import cInternalControl

objControl= cInternalControl()
dir_excel=objControl.excel_dir+objControl.excel_file

def main():
    print('Starting process, please stand by...')
    dexist=False
    dexist=os.path.exists(objControl.excel_dir)
    if dexist==False:
        os.mkdir(objControl.excel_dir)
    wb = Workbook()
    ws = wb.active
    title=objControl.excel_sheet
    ws.title = title
    lsFields=[]
    #Get cassandra columns
    keyspace=objControl.keyspace
    table=objControl.table
    columns_list=[]
    if objControl.todosCampos:
        query="select column_name from system_schema.columns WHERE keyspace_name = '"+keyspace+"' AND table_name = '"+table+"';"
        columns_list=bd.getShortQuery(query)
    else:
        #Customize fields, this option works when ALL FIELDS ARE NOT WANTED
        columns_list.append('cip')
        columns_list.append('cpc')
        columns_list.append('id')
        columns_list.append('sample')
        columns_list.append('year')

    coln=1
    for col in columns_list:
        #Write(row,column)
        #Headers (h1,...)
        h1 = ws.cell(row = 1, column = coln)
        if objControl.todosCampos:
            h1.value = col[0]
            lsFields.append(col[0])
        else:
            h1.value = col
            lsFields.append(col)
        coln=coln+1
        wb.save(dir_excel) 
    #Reading list of fields into commas for the next query
    fieldsForQuery=','.join(lsFields)    
    #Starts  the reading of periods    
    flag=os.path.isfile(dir_excel)
    if flag:
        #Expedient xls already exists
        print('Printing excel... ')
        if objControl.iterar:
            lsCondicion=['2015','2016','2017','2018','2019','2020']    
            for condicion in lsCondicion:
                query="select "+fieldsForQuery+"  from "+keyspace+"."+table+" where period_number="+str(condicion)+" ALLOW FILTERING "
                bd.getLargeQueryAndPrintToExcel_Special(query,dir_excel,title) 
        else:
            query="select "+fieldsForQuery+" from "+keyspace+"."+table+" where period_number=10 ALLOW FILTERING "
            bd.getLargeQueryAndPrintToExcel(query,dir_excel,title)



    print('The excel is ready!')        
                         
            
                    

if __name__=='__main__':
    main()    
   